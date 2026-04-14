#!/usr/bin/env python3
"""
Import books into Ghostwriter from an Excel file or Google Sheet.

Usage:
    python scripts/import_books.py --source books.xlsx
    python scripts/import_books.py --source "https://docs.google.com/spreadsheets/d/<id>"

Input format (first row must be column headers):

    title  |  notes_before_outline
    -------|----------------------
    My Book|  Write for a business audience. 8 chapters. Concise tone.

Column names are matched case-insensitively and spaces are normalised to
underscores, so "Title", "TITLE", and "Notes Before Outline" all work.

Re-running the script against the same source is safe — rows already imported
are identified by a source_id and silently skipped.
"""

import argparse
import json
import sys
from pathlib import Path

# ── Make `app` importable when running from any working directory ──
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

# Load .env from the project root before pydantic-settings instantiates.
from dotenv import load_dotenv  # noqa: E402  (after sys.path manipulation)

load_dotenv(Path(__file__).resolve().parent.parent / ".env")

from app.core.config import settings  # noqa: E402
from app.core.database import get_client  # noqa: E402

# ── Source readers ──


def _normalise_key(k: str) -> str:
    """'Notes Before Outline' → 'notes_before_outline'"""
    return str(k).strip().lower().replace(" ", "_")


def _rows_from_excel(path_str: str) -> tuple[list[dict], str]:
    """
    Read an xlsx/xls file and return (rows, source_prefix).

    *rows* is a list of dicts with normalised keys plus a private '_row' key
    holding the 1-based Excel row number.
    *source_prefix* is used to build stable source_ids.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        _die("openpyxl is required: pip install openpyxl")

    path = Path(path_str)
    if not path.exists():
        _die(f"File not found: {path_str}")

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not all_rows:
        return [], path.name

    headers = [_normalise_key(h) if h is not None else "" for h in all_rows[0]]

    result: list[dict] = []
    for excel_row_idx, row in enumerate(all_rows[1:], start=2):
        record = {
            headers[j]: (row[j] if j < len(row) else None) for j in range(len(headers))
        }
        record["_row"] = excel_row_idx
        result.append(record)

    return result, path.name


def _rows_from_sheets(url: str) -> tuple[list[dict], str]:
    """
    Read a Google Sheet and return (rows, source_prefix).

    Requires GOOGLE_SHEETS_CREDENTIALS_JSON in .env — either a file path to a
    service account JSON or the raw JSON content as a string.
    """
    try:
        import gspread
    except ImportError:
        _die("gspread is required: pip install gspread google-auth")

    creds_value = settings.google_sheets_credentials_json
    if not creds_value:
        _die(
            "GOOGLE_SHEETS_CREDENTIALS_JSON is not set in .env.\n"
            "Set it to the path of your service account JSON file, or paste "
            "the JSON content directly."
        )

    try:
        creds_path = Path(creds_value)
        if creds_path.exists():
            gc = gspread.service_account(filename=str(creds_path))
        else:
            gc = gspread.service_account_from_dict(json.loads(creds_value))
    except Exception as exc:
        _die(f"Failed to authenticate with Google: {exc}")

    try:
        sh = gc.open_by_url(url)
        ws = sh.sheet1
        records = ws.get_all_records()
    except Exception as exc:
        _die(f"Failed to read Google Sheet: {exc}")

    result: list[dict] = []
    for sheet_row_idx, record in enumerate(records, start=2):
        normalised = {_normalise_key(k): v for k, v in record.items()}
        normalised["_row"] = sheet_row_idx
        result.append(normalised)

    return result, sh.id


# ── Importer ──


def _import(rows: list[dict], source_prefix: str) -> None:
    """
    Insert each row as a book record, skipping duplicates by source_id.

    Prints a line per row and a summary at the end.
    """
    db = get_client()
    imported = skipped = errors = 0

    for record in rows:
        row_num = record.get("_row", "?")
        title = str(record.get("title") or "").strip()
        notes = str(record.get("notes_before_outline") or "").strip()

        if not title:
            print(f"  row {row_num}: skipped — 'title' is empty")
            skipped += 1
            continue

        if not notes:
            print(f"  row {row_num}: skipped — 'notes_before_outline' is empty")
            skipped += 1
            continue

        source_id = f"{source_prefix}:row{row_num}"

        existing = (
            db.table("books").select("id, title").eq("source_id", source_id).execute()
        )
        if existing.data:
            print(
                f"  row {row_num}: skipped — already imported "
                f'as "{existing.data[0]["title"]}"'
            )
            skipped += 1
            continue

        try:
            db.table("books").insert(
                {
                    "source_id": source_id,
                    "title": title,
                    "notes_before_outline": notes,
                    "outline_status": "pending",
                    "final_status": "pending",
                }
            ).execute()
            print(f'  row {row_num}: imported "{title}"')
            imported += 1
        except Exception as exc:
            print(f"  row {row_num}: error — {exc}")
            errors += 1

    print(f"\nDone. {imported} imported, {skipped} skipped, {errors} errors.")


# ── CLI ──


def _die(msg: str) -> None:
    print(f"Error: {msg}", file=sys.stderr)
    sys.exit(1)


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Import books into Ghostwriter from Excel or Google Sheets.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument(
        "--source",
        required=True,
        metavar="FILE_OR_URL",
        help="Path to an .xlsx file, or a Google Sheets URL.",
    )
    args = parser.parse_args()
    source = args.source.strip()

    if source.startswith("https://docs.google.com/spreadsheets/"):
        print(f"Source: Google Sheet")
        rows, prefix = _rows_from_sheets(source)
    else:
        print(f"Source: {source}")
        rows, prefix = _rows_from_excel(source)

    if not rows:
        print("No data rows found in source (is the file empty?)")
        return

    print(f"Found {len(rows)} data row(s). Importing…\n")
    _import(rows, prefix)


if __name__ == "__main__":
    main()
